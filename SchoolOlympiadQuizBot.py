import asyncio
import os
import logging
import sqlite3
import tempfile
import zipfile
import shutil
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ContextTypes, ConversationHandler, filters, PicklePersistence
)
import openpyxl

# Тексты кнопок
BTN_START = "Начать"
BTN_BACK_TO_YEAR = "К выбору года"
BTN_BACK_TO_EXERCISES = "К выбору задач"
BTN_BACK_TO_TASK = "↩️ К задаче"
BTN_HINT = "Подсказка"
BTN_ANSWER = "Ответ"
BTN_TOPIC_EXERCISES = "Задачи по теме"

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния
(CHOOSE_YEAR, CHOOSE_EXERCISE, CHOOSE_TOPIC_EXERCISE, TASK, HINT, ANSWER,
 ADMIN_MENU, ADMIN_UPLOAD_REPLACE, ADMIN_UPLOAD_APPEND, ADMIN_CONFIRM_CLEAR) = range(10)

# Папка для изображений
IMAGE_DIR = "images"
os.makedirs(IMAGE_DIR, exist_ok=True)


def chunks(lst, n):
    return [lst[i:i + n] for i in range(0, len(lst), n)]


class QuizBot:
    def __init__(self, admin_ids):
        self.db_path = 'quiz_bot.db'
        self.admin_ids = admin_ids
        self.init_database()

    def _get_connection(self):
        """Возвращает соединение с БД с включёнными внешними ключами."""
        conn = sqlite3.connect(self.db_path)
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def init_database(self):
        conn = self._get_connection()
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS years
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, year INTEGER UNIQUE)''')
        c.execute('''CREATE TABLE IF NOT EXISTS topics
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE)''')
        c.execute('''CREATE TABLE IF NOT EXISTS olympiads
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      year_id INTEGER,
                      excercise INTEGER,
                      task TEXT,
                      task_picture TEXT,
                      hint TEXT,
                      hint_picture TEXT,
                      answer TEXT,
                      answer_picture TEXT,
                      FOREIGN KEY (year_id) REFERENCES years (id) ON DELETE CASCADE,
                      UNIQUE(year_id, excercise))''')
        c.execute('''CREATE TABLE IF NOT EXISTS olympiad_topics
                     (olympiad_id INTEGER,
                      topic_id INTEGER,
                      FOREIGN KEY (olympiad_id) REFERENCES olympiads (id) ON DELETE CASCADE,
                      FOREIGN KEY (topic_id) REFERENCES topics (id) ON DELETE CASCADE,
                      UNIQUE(olympiad_id, topic_id))''')
        c.execute('''CREATE TABLE IF NOT EXISTS users
                     (id INTEGER PRIMARY KEY, first_name TEXT, username TEXT)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_olympiad_topics_topic
                     ON olympiad_topics(topic_id)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_olympiad_topics_olympiad
                     ON olympiad_topics(olympiad_id)''')
        conn.commit()
        conn.close()

    def save_user_to_db(self, user):
        conn = self._get_connection()
        c = conn.cursor()
        c.execute('''INSERT OR IGNORE INTO users (id, first_name, username)
                     VALUES (?, ?, ?)''', (user.id, user.first_name, user.username))
        conn.commit()
        conn.close()

    def _clean_value(self, val):
        if val is None or str(val).strip().lower() == "none" or str(val).strip() == "":
            return ""
        return str(val).strip()

    def parse_excel_and_images(self, excel_path, image_dir, replace=True):
        try:
            logger.info(f"Parsing Excel: {excel_path}, replace={replace}")
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active

            if sheet.max_row < 2:
                logger.warning("Excel file is empty")
                return False

            headers = [cell.value for cell in sheet[1]]
            logger.info(f"Headers: {headers}")

            required = ['Year', 'Excercise', 'Topic', 'Task', 'Hint', 'Answer']
            for col in required:
                if col not in headers:
                    raise ValueError(f"Missing column: {col}")
            idx = {col: headers.index(col) for col in required}

            pic_cols = {}
            for col in ['Task_picture', 'Hint_picture', 'Answer_picture']:
                if col in headers:
                    pic_cols[col] = headers.index(col)

            conn = self._get_connection()
            c = conn.cursor()
            if replace:
                c.execute("DELETE FROM olympiad_topics")
                c.execute("DELETE FROM olympiads")
                c.execute("DELETE FROM topics")
                c.execute("DELETE FROM years")
                logger.info("Cleared DB")

            inserted_years = set()
            inserted_topics = set()
            inserted = 0
            skipped = 0

            for row_num in range(2, sheet.max_row + 1):
                row = [sheet.cell(row=row_num, column=i + 1).value for i in range(len(headers))]
                if not any(str(cell).strip() if cell else '' for cell in row):
                    continue

                try:
                    year = int(float(row[idx['Year']]))
                except (ValueError, TypeError):
                    skipped += 1
                    logger.warning(f"Invalid year in row {row_num}")
                    continue

                try:
                    excercise = int(float(row[idx['Excercise']]))
                except (ValueError, TypeError):
                    skipped += 1
                    logger.warning(f"Invalid excercise in row {row_num}")
                    continue

                topic_raw = row[idx['Topic']]
                task = row[idx['Task']]
                hint = row[idx['Hint']]
                answer = row[idx['Answer']]

                t_pic = str(row[pic_cols['Task_picture']]).strip() if 'Task_picture' in pic_cols and row[pic_cols['Task_picture']] else None
                h_pic = str(row[pic_cols['Hint_picture']]).strip() if 'Hint_picture' in pic_cols and row[pic_cols['Hint_picture']] else None
                a_pic = str(row[pic_cols['Answer_picture']]).strip() if 'Answer_picture' in pic_cols and row[pic_cols['Answer_picture']] else None

                if not (year and excercise and topic_raw and (task or t_pic) and (hint or h_pic) and (answer or a_pic)):
                    skipped += 1
                    logger.warning(f"Missing data in row {row_num}")
                    continue

                topics_list = [t.strip() for t in str(topic_raw).split(',') if t.strip()]
                if not topics_list:
                    skipped += 1
                    logger.warning(f"No valid topics in row {row_num}")
                    continue

                task = self._clean_value(task)
                hint = self._clean_value(hint)
                answer = self._clean_value(answer)

                for pic in [t_pic, h_pic, a_pic]:
                    if pic and not os.path.exists(os.path.join(image_dir, pic)):
                        logger.warning(f"Picture file not found: {pic}")

                c.execute("INSERT OR IGNORE INTO years (year) VALUES (?)", (year,))
                if year not in inserted_years:
                    inserted_years.add(year)
                    logger.info(f"Added year: {year}")

                c.execute("SELECT id FROM years WHERE year = ?", (year,))
                year_id = c.fetchone()[0]

                for topic_name in topics_list:
                    c.execute("INSERT OR IGNORE INTO topics (name) VALUES (?)", (topic_name,))
                    if topic_name not in inserted_topics:
                        inserted_topics.add(topic_name)
                        logger.info(f"Added topic: {topic_name}")

                c.execute('''INSERT OR REPLACE INTO olympiads
                             (year_id, excercise, task, task_picture,
                              hint, hint_picture, answer, answer_picture)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                          (year_id, excercise, task, t_pic, hint, h_pic, answer, a_pic))

                olympiad_id = c.lastrowid

                for topic_name in topics_list:
                    c.execute("SELECT id FROM topics WHERE name = ?", (topic_name,))
                    topic_id = c.fetchone()[0]
                    c.execute('''INSERT OR IGNORE INTO olympiad_topics
                                 (olympiad_id, topic_id) VALUES (?, ?)''',
                              (olympiad_id, topic_id))

                inserted += 1

            conn.commit()
            conn.close()
            logger.info(f"Loaded: {inserted} exercises, {len(inserted_topics)} topics, "
                        f"{len(inserted_years)} years, skipped: {skipped}")
            return True

        except Exception as e:
            logger.error(f"Error parsing Excel: {e}", exc_info=True)
            return False

    def clear_database(self):
        try:
            conn = self._get_connection()
            c = conn.cursor()
            c.execute("DELETE FROM olympiads")
            c.execute("DELETE FROM topics")
            c.execute("DELETE FROM years")
            conn.commit()
            conn.close()
            logger.info("Database cleared successfully.")
        except Exception as e:
            logger.error(f"Error clearing database: {e}")

    def clear_images(self):
        try:
            if os.path.exists(IMAGE_DIR):
                for filename in os.listdir(IMAGE_DIR):
                    file_path = os.path.join(IMAGE_DIR, filename)
                    try:
                        if os.path.isfile(file_path):
                            os.unlink(file_path)
                            logger.info(f"Deleted image file: {filename}")
                    except Exception as e:
                        logger.error(f"Failed to delete {file_path}: {e}")
            logger.info("All image files deleted.")
        except Exception as e:
            logger.error(f"Error in clear_images: {e}")

    def get_years_from_db(self):
        conn = self._get_connection()
        c = conn.cursor()
        c.execute("SELECT year FROM years ORDER BY year")
        years = [row[0] for row in c.fetchall()]
        conn.close()
        return years

    def get_exercises_for_year(self, year):
        conn = self._get_connection()
        c = conn.cursor()
        c.execute('''SELECT o.excercise
                     FROM olympiads o
                     JOIN years y ON o.year_id = y.id
                     WHERE y.year = ?
                     ORDER BY o.excercise''', (year,))
        exercises = [{'excercise': row[0]} for row in c.fetchall()]
        conn.close()
        return exercises

    def get_all_exercises_by_topics_with_matching_topics(self, topic_list):
        """Возвращает все задания по указанным темам из всех годов."""
        conn = self._get_connection()
        c = conn.cursor()
        placeholders = ','.join('?' * len(topic_list))
        query = f'''
            SELECT DISTINCT y.year, o.excercise, GROUP_CONCAT(t.name, ', ') as matching_topics
            FROM olympiads o
            JOIN years y ON o.year_id = y.id
            JOIN olympiad_topics ot ON o.id = ot.olympiad_id
            JOIN topics t ON ot.topic_id = t.id
            WHERE t.name IN ({placeholders})
            GROUP BY y.year, o.excercise
            ORDER BY y.year, o.excercise
        '''
        c.execute(query, topic_list)
        results = [
            {'year': row[0], 'excercise': row[1], 'matching_topics': row[2]}
            for row in c.fetchall()
        ]
        conn.close()
        return results

    def get_task_for_year_and_exercise(self, year, excercise):
        """Возвращает задание с темами или None если не найдено."""
        conn = self._get_connection()
        c = conn.cursor()
        c.execute('''SELECT o.id, o.excercise, o.task, o.task_picture,
                            o.hint, o.hint_picture,
                            o.answer, o.answer_picture
                     FROM olympiads o
                     JOIN years y ON o.year_id = y.id
                     WHERE y.year = ? AND o.excercise = ?''', (year, excercise))
        row = c.fetchone()
        if not row:
            conn.close()
            return None

        task = {
            'id': row[0],
            'excercise': row[1],
            'task': row[2],
            't_pic': row[3],
            'hint': row[4],
            'h_pic': row[5],
            'answer': row[6],
            'a_pic': row[7],
        }

        c.execute('''SELECT t.name
                     FROM topics t
                     JOIN olympiad_topics ot ON t.id = ot.topic_id
                     WHERE ot.olympiad_id = ?''', (task['id'],))
        task['topics'] = [r[0] for r in c.fetchall()]

        conn.close()
        return task

    def get_exercises_by_topics_and_year(self, year, topic_list):
        conn = self._get_connection()
        c = conn.cursor()
        placeholders = ','.join('?' * len(topic_list))
        query = f'''
            SELECT DISTINCT o.excercise
            FROM olympiads o
            JOIN years y ON o.year_id = y.id
            JOIN olympiad_topics ot ON o.id = ot.olympiad_id
            JOIN topics t ON ot.topic_id = t.id
            WHERE y.year = ? AND t.name IN ({placeholders})
            ORDER BY o.excercise
        '''
        c.execute(query, [year] + topic_list)
        exercises = [{'excercise': row[0]} for row in c.fetchall()]
        conn.close()
        return exercises

    # === Helpers ===

    def _build_year_keyboard(self, years):
        return chunks([str(y) for y in years], 4)

    def _build_exercise_keyboard(self, year, exercises):
        buttons = [f"{year} задание {ex['excercise']}" for ex in exercises]
        keyboard = chunks(buttons, 3)
        keyboard.append([BTN_BACK_TO_YEAR])
        return keyboard

    # === Handlers ===

    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        self.save_user_to_db(user)
        years = self.get_years_from_db()

        keyboard = [[BTN_START]]
        if user.id in self.admin_ids:
            keyboard.append(["🛡️ Админка"])

        await update.message.reply_text(
            f"Привет! Нажмите «{BTN_START}», чтобы выбрать год.",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

        if not years and user.id not in self.admin_ids:
            await update.message.reply_text("Нет доступных годов. Обратитесь к админу.")
            return ConversationHandler.END

        return CHOOSE_YEAR

    async def choose_year(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text

        # BTN_START и BTN_BACK_TO_YEAR — показать список годов
        if text in (BTN_START, BTN_BACK_TO_YEAR):
            years = self.get_years_from_db()
            if not years:
                await update.message.reply_text("Нет доступных годов.")
                return ConversationHandler.END
            await update.message.reply_text(
                "Выберите год:",
                reply_markup=ReplyKeyboardMarkup(self._build_year_keyboard(years),
                                                 one_time_keyboard=False)
            )
            return CHOOSE_YEAR

        try:
            year = int(text)
        except ValueError:
            years = self.get_years_from_db()
            if not years:
                await update.message.reply_text("Нет доступных годов.")
                return ConversationHandler.END
            await update.message.reply_text(
                "Выберите год из списка.",
                reply_markup=ReplyKeyboardMarkup(self._build_year_keyboard(years),
                                                 one_time_keyboard=False)
            )
            return CHOOSE_YEAR

        exercises = self.get_exercises_for_year(year)
        if not exercises:
            await update.message.reply_text("В этом году нет заданий.")
            return CHOOSE_YEAR

        await update.message.reply_text(
            f"Выберите задание для {year} года:",
            reply_markup=ReplyKeyboardMarkup(self._build_exercise_keyboard(year, exercises),
                                             one_time_keyboard=False)
        )
        context.user_data['state'] = {'year': year, 'exercises': exercises}
        return CHOOSE_EXERCISE

    async def choose_exercise(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state:
            return await self.start(update, context)

        text = update.message.text
        year = state['year']

        try:
            if f"{year} задание " in text:
                excercise = int(text.split()[-1])
            else:
                raise ValueError("Неверный формат")
        except (ValueError, IndexError):
            await update.message.reply_text("Пожалуйста, выберите задание из списка.")
            return CHOOSE_EXERCISE

        task = self.get_task_for_year_and_exercise(year, excercise)
        if not task:
            await update.message.reply_text("Задание не найдено.")
            return CHOOSE_EXERCISE

        topics = task['topics']
        context.user_data['state'] = {
            'year': year,
            'exercises': state['exercises'],
            'current_task': task,
            'current_topics': topics,
            'current_topic_str': ", ".join(topics) if topics else "Без темы",
        }

        await self.show_task(update, task)
        return TASK

    async def choose_topic_exercise(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state:
            return await self.start(update, context)

        text = update.message.text

        # Формат: "2023 задание 5 (Алгебра, Геометрия)"
        try:
            clean_text = text.split(' (')[0]
            parts = clean_text.split()
            if len(parts) >= 3 and parts[1] == "задание":
                year = int(parts[0])
                excercise = int(parts[2])
            else:
                raise ValueError("Неверный формат")
        except (ValueError, IndexError):
            await update.message.reply_text("Пожалуйста, выберите задание из списка.")
            return CHOOSE_TOPIC_EXERCISE

        task = self.get_task_for_year_and_exercise(year, excercise)
        if not task:
            await update.message.reply_text("Задание не найдено.")
            return CHOOSE_TOPIC_EXERCISE

        topics = task['topics']
        context.user_data['state'] = {
            'year': year,
            'exercises': self.get_exercises_for_year(year),
            'current_task': task,
            'current_topics': topics,
            'current_topic_str': ", ".join(topics) if topics else "Без темы",
        }

        await self.show_task(update, task)
        return TASK

    async def show_task(self, update: Update, q):
        task_text = q['task'] or ""
        if task_text:
            await update.message.reply_text(f"❓ Задача:\n{task_text}")
        elif not q['t_pic']:
            await update.message.reply_text("❓ Задача: (текст не указан)")

        if q['t_pic']:
            pic_path = os.path.join(IMAGE_DIR, q['t_pic'])
            if os.path.exists(pic_path):
                await update.message.reply_photo(photo=pic_path)
            else:
                await update.message.reply_text(f"🖼️ Изображение задачи не найдено: {q['t_pic']}")

        keyboard = [
            [BTN_HINT, BTN_ANSWER],
            [BTN_TOPIC_EXERCISES],
            [BTN_BACK_TO_EXERCISES, BTN_BACK_TO_YEAR],
        ]
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=False)
        )

    async def show_hint(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state or 'current_task' not in state:
            await update.message.reply_text("Нет активной задачи.")
            return CHOOSE_YEAR

        q = state['current_task']
        hint_text = q['hint'] or ""

        if hint_text:
            await update.message.reply_text(f"💡 Подсказка:\n{hint_text}")
        elif not q['h_pic']:
            await update.message.reply_text("💡 Подсказка не указана.")

        if q['h_pic']:
            pic_path = os.path.join(IMAGE_DIR, q['h_pic'])
            if os.path.exists(pic_path):
                await update.message.reply_photo(photo=pic_path)
            else:
                await update.message.reply_text(f"🖼️ Изображение подсказки не найдено: {q['h_pic']}")

        keyboard = [
            [BTN_ANSWER, BTN_BACK_TO_TASK],
            [BTN_TOPIC_EXERCISES],
            [BTN_BACK_TO_EXERCISES, BTN_BACK_TO_YEAR],
        ]
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=False)
        )
        return HINT

    async def show_answer(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state or 'current_task' not in state:
            await update.message.reply_text("Нет активной задачи.")
            return CHOOSE_YEAR

        q = state['current_task']
        answer_text = q['answer'] or ""

        if answer_text:
            await update.message.reply_text(f"✅ Ответ:\n{answer_text}")
        elif not q['a_pic']:
            await update.message.reply_text("✅ Ответ не указан.")

        if q['a_pic']:
            pic_path = os.path.join(IMAGE_DIR, q['a_pic'])
            if os.path.exists(pic_path):
                await update.message.reply_photo(photo=pic_path)
            else:
                await update.message.reply_text(f"🖼️ Изображение ответа не найдено: {q['a_pic']}")

        keyboard = [
            [BTN_TOPIC_EXERCISES],
            [BTN_BACK_TO_EXERCISES, BTN_BACK_TO_YEAR],
        ]
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=False)
        )
        return ANSWER

    async def show_topic_exercises(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state or 'current_topics' not in state:
            return await self.start(update, context)

        topics = state['current_topics']
        current_task = state.get('current_task')
        current_year = state.get('year')

        exercises_data = self.get_all_exercises_by_topics_with_matching_topics(topics)

        # Исключаем текущее задание из списка
        if current_task:
            exercises_data = [
                d for d in exercises_data
                if not (d['year'] == current_year and d['excercise'] == current_task['excercise'])
            ]

        if not exercises_data:
            await update.message.reply_text("Нет других заданий по этим темам.")
            return await self.show_task_from_state(update, context)

        buttons = [
            f"{d['year']} задание {d['excercise']} ({d['matching_topics']})"
            for d in exercises_data
        ]
        keyboard = chunks(buttons, 2)
        keyboard.append([BTN_BACK_TO_EXERCISES, BTN_BACK_TO_YEAR])

        await update.message.reply_text(
            f"Задания по темам {', '.join(topics)} по всем годам:",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=False)
        )
        return CHOOSE_TOPIC_EXERCISE

    async def show_task_from_state(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state or 'current_task' not in state:
            return await self.start(update, context)
        await self.show_task(update, state['current_task'])
        return TASK

    async def back_to_year_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data.pop('state', None)
        years = self.get_years_from_db()
        if not years:
            await update.message.reply_text("Нет доступных годов.")
            return ConversationHandler.END
        await update.message.reply_text(
            "Выберите год:",
            reply_markup=ReplyKeyboardMarkup(self._build_year_keyboard(years),
                                             one_time_keyboard=False)
        )
        return CHOOSE_YEAR

    async def back_to_exercises(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        state = context.user_data.get('state')
        if not state or 'year' not in state or 'exercises' not in state:
            return await self.start(update, context)

        year = state['year']
        exercises = state['exercises']
        await update.message.reply_text(
            f"Выберите задание для {year} года:",
            reply_markup=ReplyKeyboardMarkup(self._build_exercise_keyboard(year, exercises),
                                             one_time_keyboard=False)
        )
        return CHOOSE_EXERCISE

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data.pop('state', None)
        await update.message.reply_text("Операция отменена.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    # === Admin handlers ===

    async def admin_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        logger.info(f"Попытка доступа к админке от пользователя ID={user_id}, "
                    f"админы={self.admin_ids}")

        if user_id not in self.admin_ids:
            await update.message.reply_text(f"❌ Доступ запрещён. Ваш ID: {user_id}")
            return ConversationHandler.END

        await self._show_admin_menu(update)
        return ADMIN_MENU

    async def _show_admin_menu(self, update: Update):
        keyboard = [
            ['📁 Загрузить данные', '📥 Дополнить данные'],
            ['🧹 Удалить данные', '↩️ Выйти'],
        ]
        await update.message.reply_text(
            "🛡️ Админ-панель:\n"
            "• 📁 — заменить все данные\n"
            "• 📥 — добавить к существующим\n"
            "• 🧹 — удалить всё",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

    async def admin_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if user_id not in self.admin_ids:
            await update.message.reply_text("❌ Доступ запрещён.", reply_markup=ReplyKeyboardRemove())
            return ConversationHandler.END

        choice = update.message.text
        if choice == "↩️ Выйти":
            await update.message.reply_text("Вы вышли.", reply_markup=ReplyKeyboardRemove())
            return ConversationHandler.END
        elif choice == "📁 Загрузить данные":
            await update.message.reply_text("Отправьте ZIP-архив с Excel и изображениями.")
            return ADMIN_UPLOAD_REPLACE
        elif choice == "📥 Дополнить данные":
            await update.message.reply_text(
                "Отправьте ZIP-архив с Excel и изображениями для дополнения.")
            return ADMIN_UPLOAD_APPEND
        elif choice == "🧹 Удалить данные":
            await update.message.reply_text(
                "Точно удалить все данные?",
                reply_markup=ReplyKeyboardMarkup([['✅ Да', '❌ Нет']], resize_keyboard=True)
            )
            return ADMIN_CONFIRM_CLEAR
        else:
            await update.message.reply_text("Выберите действие из меню.")
            return ADMIN_MENU

    async def admin_upload_file(self, update: Update, context: ContextTypes.DEFAULT_TYPE,
                                replace=True):
        if not update.message.document:
            await update.message.reply_text("Отправьте ZIP-архив.")
            return ADMIN_UPLOAD_REPLACE if replace else ADMIN_UPLOAD_APPEND

        if not update.message.document.file_name.lower().endswith('.zip'):
            await update.message.reply_text("Файл должен быть ZIP-архивом (.zip).")
            return ADMIN_UPLOAD_REPLACE if replace else ADMIN_UPLOAD_APPEND

        file = await update.message.document.get_file()
        with tempfile.TemporaryDirectory() as tmp_dir:
            zip_path = os.path.join(tmp_dir, "data.zip")
            await file.download_to_drive(zip_path)

            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(tmp_dir)
            except zipfile.BadZipFile:
                await update.message.reply_text("Неверный ZIP-файл.")
                return ADMIN_UPLOAD_REPLACE if replace else ADMIN_UPLOAD_APPEND

            excel_files = [f for f in os.listdir(tmp_dir) if f.endswith(('.xlsx', '.xls'))]
            if not excel_files:
                await update.message.reply_text("В архиве нет Excel-файла.")
                return ADMIN_UPLOAD_REPLACE if replace else ADMIN_UPLOAD_APPEND

            excel_path = os.path.join(tmp_dir, excel_files[0])

            for item in os.listdir(tmp_dir):
                if item.lower().endswith(('.jpg', '.jpeg', '.png')):
                    shutil.copy(os.path.join(tmp_dir, item), os.path.join(IMAGE_DIR, item))

            success = self.parse_excel_and_images(excel_path, tmp_dir, replace=replace)

        if success:
            await update.message.reply_text("✅ Данные успешно загружены!")
            await self._show_admin_menu(update)
            return ADMIN_MENU
        else:
            await update.message.reply_text("❌ Ошибка при загрузке данных.")
            return ADMIN_UPLOAD_REPLACE if replace else ADMIN_UPLOAD_APPEND

    async def admin_confirm_clear(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.message.text == "✅ Да":
            self.clear_database()
            self.clear_images()
            years = self.get_years_from_db()
            images_count = len(os.listdir(IMAGE_DIR)) if os.path.exists(IMAGE_DIR) else 0
            await update.message.reply_text(
                f"🧹 Все данные удалены.\n"
                f"Осталось лет в БД: {len(years)}\n"
                f"Осталось изображений: {images_count}"
            )
        else:
            await update.message.reply_text("Удаление отменено.")

        await self._show_admin_menu(update)
        return ADMIN_MENU


# === Main ===

async def main():
    TOKEN = os.getenv("BOT_TOKEN")
    if not TOKEN:
        raise ValueError("BOT_TOKEN не задан!")

    admin_ids_str = os.getenv("ADMIN_IDS", "").strip()
    if not admin_ids_str:
        raise ValueError("ADMIN_IDS не задан! Укажите ID администраторов через запятую")

    admin_ids = [int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()]
    if not admin_ids:
        raise ValueError("ADMIN_IDS содержит некорректные ID")

    logger.info(f"Администраторы: {admin_ids}")
    quiz_bot = QuizBot(admin_ids=admin_ids)
    persistence = PicklePersistence(filepath="conversation_states.pkl")
    app = Application.builder().token(TOKEN).persistence(persistence).build()

    # admin_handler регистрируется первым, чтобы "🛡️ Админка" перехватывалась
    # раньше conv_handler в любом состоянии основного диалога
    admin_handler = ConversationHandler(
        entry_points=[
            CommandHandler('admin', quiz_bot.admin_start),
            MessageHandler(filters.Text(["🛡️ Админка"]), quiz_bot.admin_start),
        ],
        states={
            ADMIN_MENU: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, quiz_bot.admin_menu)
            ],
            ADMIN_UPLOAD_REPLACE: [
                MessageHandler(filters.Document.ALL, quiz_bot.admin_upload_file),
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               lambda u, c: quiz_bot.admin_upload_file(u, c, replace=True)),
            ],
            ADMIN_UPLOAD_APPEND: [
                MessageHandler(filters.Document.ALL,
                               lambda u, c: quiz_bot.admin_upload_file(u, c, replace=False)),
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               lambda u, c: quiz_bot.admin_upload_file(u, c, replace=False)),
            ],
            ADMIN_CONFIRM_CLEAR: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, quiz_bot.admin_confirm_clear)
            ],
        },
        fallbacks=[
            CommandHandler('cancel', quiz_bot.cancel),
            MessageHandler(filters.Text(["Отмена", "Cancel"]), quiz_bot.cancel),
        ],
        name="admin_conv",
        persistent=True,
        allow_reentry=True,
    )

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', quiz_bot.start)],
        states={
            CHOOSE_YEAR: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, quiz_bot.choose_year),
            ],
            CHOOSE_EXERCISE: [
                MessageHandler(filters.Text([BTN_BACK_TO_YEAR]),
                               quiz_bot.back_to_year_selection),
                MessageHandler(filters.TEXT & ~filters.COMMAND, quiz_bot.choose_exercise),
            ],
            CHOOSE_TOPIC_EXERCISE: [
                MessageHandler(filters.Text([BTN_BACK_TO_YEAR]),
                               quiz_bot.back_to_year_selection),
                MessageHandler(filters.Text([BTN_BACK_TO_EXERCISES]),
                               quiz_bot.back_to_exercises),
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               quiz_bot.choose_topic_exercise),
            ],
            TASK: [
                MessageHandler(filters.Text([BTN_HINT]), quiz_bot.show_hint),
                MessageHandler(filters.Text([BTN_ANSWER]), quiz_bot.show_answer),
                MessageHandler(filters.Text([BTN_TOPIC_EXERCISES]),
                               quiz_bot.show_topic_exercises),
                MessageHandler(filters.Text([BTN_BACK_TO_EXERCISES]),
                               quiz_bot.back_to_exercises),
                MessageHandler(filters.Text([BTN_BACK_TO_YEAR]),
                               quiz_bot.back_to_year_selection),
            ],
            HINT: [
                MessageHandler(filters.Text([BTN_ANSWER]), quiz_bot.show_answer),
                MessageHandler(filters.Text([BTN_BACK_TO_TASK]),
                               quiz_bot.show_task_from_state),
                MessageHandler(filters.Text([BTN_TOPIC_EXERCISES]),
                               quiz_bot.show_topic_exercises),
                MessageHandler(filters.Text([BTN_BACK_TO_EXERCISES]),
                               quiz_bot.back_to_exercises),
                MessageHandler(filters.Text([BTN_BACK_TO_YEAR]),
                               quiz_bot.back_to_year_selection),
            ],
            ANSWER: [
                MessageHandler(filters.Text([BTN_TOPIC_EXERCISES]),
                               quiz_bot.show_topic_exercises),
                MessageHandler(filters.Text([BTN_BACK_TO_EXERCISES]),
                               quiz_bot.back_to_exercises),
                MessageHandler(filters.Text([BTN_BACK_TO_YEAR]),
                               quiz_bot.back_to_year_selection),
            ],
        },
        fallbacks=[
            CommandHandler('cancel', quiz_bot.cancel),
            CommandHandler('start', quiz_bot.start),
            MessageHandler(filters.Text([BTN_START]), quiz_bot.start),
            MessageHandler(filters.Text(["Отмена", "Cancel"]), quiz_bot.cancel),
        ],
        name="main_conversation",
        persistent=True,
    )

    app.add_handler(admin_handler)
    app.add_handler(conv_handler)

    logger.info("Запуск бота в режиме polling...")
    await app.initialize()
    await app.start()
    await app.updater.start_polling()
    logger.info("Бот работает.")

    try:
        while True:
            await asyncio.sleep(3600)
    except KeyboardInterrupt:
        logger.info("Остановка...")
    finally:
        await app.updater.stop()
        await app.stop()
        await app.shutdown()


if __name__ == '__main__':
    asyncio.run(main())
